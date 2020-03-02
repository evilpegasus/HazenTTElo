from elopy import *
import openpyxl

def calculate(file):
    """
    Calculate the elo ratings of matches listed on an Excel spreadsheet starting from top to bottom.
    The first row should be a header with labels for date, player 1, player 1 score, player 2, and player 2 score, in that order.
    parameters:
        file: location of the Excel file as a string
    """
    i = Implementation()
    wb = openpyxl.load_workbook(file)
    sheet = wb.worksheets[0]

    total_rows = sheet.max_row
    total_col = sheet.max_column
    print("Rows: " + str(total_rows))
    print("Cols: " + str(total_col))

    for r in range(2, total_rows + 1):
        player1_name = sheet.cell(row = r, column = 2).value
        player2_name = sheet.cell(row = r, column = 4).value
        #try to add players to the list
        #if they already exist this will do nothing
        i.addPlayer(player1_name)
        i.addPlayer(player2_name)

        player1_score = sheet.cell(row = r, column = 3).value
        player2_score = sheet.cell(row = r, column = 5).value

        if player1_score > player2_score:
            i.recordMatch(player1_name, player2_name, winner = player1_name)
            print("Recorded player 1 win")
            print(i.getRatingList())
        elif player1_score < player2_score:
            i.recordMatch(player1_name, player2_name, winner = player2_name)
            print("Recorded player 2 win")
            print(i.getRatingList())
        else:
            i.recordMatch(player1_name, player2_name, draw = True)
            print("Recorded draw")
            print(i.getRatingList())

#return list of players and elo

#display on a webpage


calculate("Log.xlsx")